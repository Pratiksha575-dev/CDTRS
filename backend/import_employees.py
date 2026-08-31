import os
import sys
import json
import argparse
from pathlib import Path
from typing import List, Dict, Any

BASE_DIR = Path(__file__).resolve().parent
JSON_SEED_PATH = BASE_DIR / 'data' / 'seed_data.json'

def parse_excel_or_csv(file_path: str) -> List[Dict[str, Any]]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f'File not found: {file_path}')

    records = []
    ext = path.suffix.lower()

    if ext in ('.xlsx', '.xls'):
        try:
            import pandas as pd
            df = pd.read_excel(file_path)
            df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
            for _, row in df.iterrows():
                if pd.isna(row.get('full_name')) and pd.isna(row.get('name')):
                    continue
                records.append({
                    'employee_code': str(row.get('employee_code') or row.get('code') or '').strip(),
                    'username': str(row.get('username') or '').strip(),
                    'full_name': str(row.get('full_name') or row.get('name') or '').strip(),
                    'department': str(row.get('department') or row.get('dept') or 'General').strip(),
                    'designation': str(row.get('designation') or row.get('title') or 'Staff').strip(),
                    'email': str(row.get('email') or '').strip(),
                    'outlook_email': str(row.get('outlook_email') or row.get('outlook') or '').strip(),
                    'gov_email': str(row.get('gov_email') or row.get('nic_email') or '').strip(),
                    'default_password': str(row.get('password') or 'cdtrs@emp').strip(),
                })
        except ImportError:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                headers = [str(c.value).strip().lower().replace(' ', '_') for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = dict(zip(headers, row))
                    name = row_dict.get('full_name') or row_dict.get('name')
                    if not name:
                        continue
                    records.append({
                        'employee_code': str(row_dict.get('employee_code') or row_dict.get('code') or '').strip(),
                        'username': str(row_dict.get('username') or '').strip(),
                        'full_name': str(name).strip(),
                        'department': str(row_dict.get('department') or row_dict.get('dept') or 'General').strip(),
                        'designation': str(row_dict.get('designation') or row_dict.get('title') or 'Staff').strip(),
                        'email': str(row_dict.get('email') or '').strip(),
                        'outlook_email': str(row_dict.get('outlook_email') or row_dict.get('outlook') or '').strip(),
                        'gov_email': str(row_dict.get('gov_email') or row_dict.get('nic_email') or '').strip(),
                        'default_password': str(row_dict.get('password') or 'cdtrs@emp').strip(),
                    })
            except ImportError:
                print('[ERROR] Please install pandas or openpyxl (pip install openpyxl pandas) to read Excel files.')
                return []

    elif ext == '.csv':
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                norm_row = {k.strip().lower().replace(' ', '_'): v for k, v in row.items() if k}
                name = norm_row.get('full_name') or norm_row.get('name')
                if not name:
                    continue
                records.append({
                    'employee_code': str(norm_row.get('employee_code') or norm_row.get('code') or '').strip(),
                    'username': str(norm_row.get('username') or '').strip(),
                    'full_name': str(name).strip(),
                    'department': str(norm_row.get('department') or norm_row.get('dept') or 'General').strip(),
                    'designation': str(norm_row.get('designation') or norm_row.get('title') or 'Staff').strip(),
                    'email': str(norm_row.get('email') or '').strip(),
                    'outlook_email': str(norm_row.get('outlook_email') or norm_row.get('outlook') or '').strip(),
                    'gov_email': str(norm_row.get('gov_email') or norm_row.get('nic_email') or '').strip(),
                    'default_password': str(norm_row.get('password') or 'cdtrs@emp').strip(),
                })

    return records


def update_seed_json(new_employees: List[Dict[str, Any]], output_path: Path = JSON_SEED_PATH):
    if output_path.exists():
        with open(output_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {'departments': [], 'system_users': [], 'employees': []}

    existing_dept_names = {d['name'].lower() for d in data.get('departments', [])}
    for emp in new_employees:
        dept_name = emp.get('department')
        if dept_name and dept_name.lower() not in existing_dept_names:
            code = dept_name[:4].upper()
            data.setdefault('departments', []).append({'name': dept_name, 'code': code})
            existing_dept_names.add(dept_name.lower())

        if not emp.get('username') and emp.get('full_name'):
            parts = emp['full_name'].lower().split()
            emp['username'] = f'emp_{parts[0]}' if len(parts) == 1 else f'emp_{parts[0]}_{parts[-1]}'

    data['employees'] = new_employees

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

    print(f'[SUCCESS] Updated {output_path} with {len(new_employees)} employee records.')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import Employee Data from Excel/CSV into CDTRS JSON Seeding')
    parser.add_argument('file', help='Path to .xlsx or .csv employee roster file')
    parser.add_argument('--seed-db', action='store_true', help='Immediately seed database after updating JSON')

    args = parser.parse_args()
    employees = parse_excel_or_csv(args.file)
    if employees:
        update_seed_json(employees)
        if args.seed_db:
            import seed
            seed.main(reset=False)
